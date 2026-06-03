import csv
from calendar import monthrange
from datetime import datetime, timedelta
from io import StringIO

from django.contrib.auth import get_user_model
from django.http import HttpResponse
from django.db import transaction
from django.db.models import Count, Max, Min, Q, Sum
from django.db.utils import OperationalError, ProgrammingError
from django.shortcuts import get_object_or_404
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from rest_framework import status, viewsets
from rest_framework.decorators import action
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response

from .models import (
    AttendanceStatus,
    CalendarDayCell,
    CalendarEmployeeAssignment,
    EmployeeVisualCategory,
    MonthlyOperationCalendar,
    OperationalCode,
    OperationalPosition,
    OperationCalendarTemplate,
    OperationCalendarTemplateAssignment,
    OperationCalendarTemplateCell,
    OperationCalendarHistory,
    PositionDailyRequirement,
    HikitsuguiOccurrenceCategory,
    HikitsuguiReport,
    HikitsuguiItem,
    ProductionMonitorSource,
    ProductionSnapshot,
    ProductionMachineStatus,
    ProductionMetrics,
    OperationsSettings,
    EmployeeAdministrativeNote,
    AttendanceTimecardRecord,
    RotationGroupStyle,
    WorkTimeCode,
    OperationRole,
    UserOperationProfile,
    UserOperationScope,
    OperationAccessAuditLog,
)
from .permissions import (
    AttendanceDashboardPermission,
    EmployeeAdminNotePermission,
    HikitsuguiPermission,
    OperationsCalendarPermission,
    OperationsMasterDataPermission,
    OperationsSettingsPermission,
    OperationsRBACManagementPermission,
    ScheduleWritePermission,
)
from .rbac import get_user_operation_permissions_payload, get_user_operation_profile, get_user_operation_role_code
from .serializers import (
    AttendanceStatusSerializer,
    CalendarDayCellSerializer,
    CalendarEmployeeAssignmentSerializer,
    EmployeeVisualCategorySerializer,
    MonthlyOperationCalendarSerializer,
    OperationalCodeSerializer,
    OperationalPositionSerializer,
    OperationCalendarTemplateSerializer,
    OperationCalendarHistorySerializer,
    PositionDailyRequirementSerializer,
    HikitsuguiOccurrenceCategorySerializer,
    HikitsuguiReportSerializer,
    HikitsuguiItemSerializer,
    ProductionMonitorSourceSerializer,
    ProductionSnapshotSerializer,
    ProductionMachineStatusSerializer,
    ProductionMetricsSerializer,
    OperationsSettingsSerializer,
    EmployeeAdministrativeNoteSerializer,
    RotationGroupStyleSerializer,
    WorkTimeCodeSerializer,
    OperationRoleSerializer,
    UserOperationProfileSerializer,
    UserOperationScopeSerializer,
    OperationAccessAuditLogSerializer,
    OperationAccessUserListSerializer,
)
from .services import (
    build_calendar_cell_parser_context,
    calculate_cell_work_minutes,
    compare_timecard_to_calendar,
    generate_calendar_schedule,
    get_assignment_sort_key,
    import_calendar_employees,
    preview_calendar_employee_candidates,
    parse_calendar_cell_value,
    recalculate_calendar_totals,
    sync_calendar_assignments_from_master,
    _normalize_employee_code,
)
from master.models import Department, Process, Shift


class ActorMixin:
    def _actor(self):
        return self.request.user if self.request.user.is_authenticated else None

    def perform_create(self, serializer):
        user = self._actor()
        serializer.save(created_by=user, updated_by=user)

    def perform_update(self, serializer):
        serializer.save(updated_by=self._actor())


def _apply_operational_scope_filter(queryset, user, *, department_field, process_field=None, shift_field=None):
    if not user or not user.is_authenticated:
        return queryset.none()
    if user.is_superuser:
        return queryset

    role_code = get_user_operation_role_code(user)
    if role_code in {"director", "vice_director", "hr"}:
        return queryset

    profile = get_user_operation_profile(user)
    if not profile:
        # Transitional compatibility: legacy accounts-role users keep current visibility.
        return queryset

    scopes = profile.scopes.filter(is_active=True)
    if not scopes.exists():
        return queryset.none()

    scope_filter = Q()
    for scope in scopes:
        clause = Q()
        if scope.department_id is not None:
            clause &= Q(**{department_field: scope.department_id})
        if process_field and scope.process_id is not None:
            clause &= Q(**{process_field: scope.process_id})
        if shift_field and scope.shift_id is not None:
            clause &= Q(**{shift_field: scope.shift_id})
        scope_filter |= clause

    if not scope_filter:
        return queryset.none()
    return queryset.filter(scope_filter)


class OperationalPositionViewSet(ActorMixin, viewsets.ModelViewSet):
    queryset = OperationalPosition.objects.select_related("department", "building_floor")
    serializer_class = OperationalPositionSerializer
    permission_classes = [OperationsMasterDataPermission]

    def get_queryset(self):
        queryset = super().get_queryset()
        department = self.request.query_params.get("department")
        if department:
            queryset = queryset.filter(department_id=department)
        return queryset


class AttendanceStatusViewSet(ActorMixin, viewsets.ModelViewSet):
    queryset = AttendanceStatus.objects.all()
    serializer_class = AttendanceStatusSerializer
    permission_classes = [OperationsMasterDataPermission]


class WorkTimeCodeViewSet(ActorMixin, viewsets.ModelViewSet):
    queryset = WorkTimeCode.objects.all()
    serializer_class = WorkTimeCodeSerializer
    permission_classes = [OperationsMasterDataPermission]


class RotationGroupStyleViewSet(ActorMixin, viewsets.ModelViewSet):
    queryset = RotationGroupStyle.objects.all()
    serializer_class = RotationGroupStyleSerializer
    permission_classes = [OperationsMasterDataPermission]


class EmployeeVisualCategoryViewSet(ActorMixin, viewsets.ModelViewSet):
    queryset = EmployeeVisualCategory.objects.all()
    serializer_class = EmployeeVisualCategorySerializer
    permission_classes = [OperationsMasterDataPermission]


class OperationalCodeViewSet(ActorMixin, viewsets.ModelViewSet):
    queryset = OperationalCode.objects.select_related("attendance_status", "work_time_code")
    serializer_class = OperationalCodeSerializer
    permission_classes = [OperationsMasterDataPermission]


class OperationCalendarTemplateViewSet(ActorMixin, viewsets.ModelViewSet):
    queryset = OperationCalendarTemplate.objects.select_related("department", "process", "shift", "created_by")
    serializer_class = OperationCalendarTemplateSerializer
    permission_classes = [OperationsCalendarPermission]

    def get_queryset(self):
        queryset = super().get_queryset()
        department = self.request.query_params.get("department")
        process = self.request.query_params.get("process")
        shift = self.request.query_params.get("shift")
        if department:
            queryset = queryset.filter(department_id=department)
        if process:
            queryset = queryset.filter(process_id=process)
        if shift:
            queryset = queryset.filter(shift_id=shift)
        return queryset


class HikitsuguiOccurrenceCategoryViewSet(ActorMixin, viewsets.ModelViewSet):
    queryset = HikitsuguiOccurrenceCategory.objects.all()
    serializer_class = HikitsuguiOccurrenceCategorySerializer
    permission_classes = [HikitsuguiPermission]


class HikitsuguiReportViewSet(ActorMixin, viewsets.ModelViewSet):
    queryset = HikitsuguiReport.objects.select_related(
        "calendar",
        "shift",
        "process",
        "responsible_employee",
        "responsible_assignment",
    ).prefetch_related("items", "items__category")
    serializer_class = HikitsuguiReportSerializer
    permission_classes = [HikitsuguiPermission]

    def get_queryset(self):
        queryset = super().get_queryset()
        params = self.request.query_params
        calendar = params.get("calendar")
        shift = params.get("shift")
        process = params.get("process")
        status_param = params.get("status")
        priority = params.get("priority")
        date_from = params.get("date_from")
        date_to = params.get("date_to")
        responsible_employee = params.get("responsible_employee")
        if calendar:
            queryset = queryset.filter(calendar_id=calendar)
        if shift:
            queryset = queryset.filter(shift_id=shift)
        if process:
            queryset = queryset.filter(process_id=process)
        if status_param:
            queryset = queryset.filter(status=status_param)
        if priority:
            queryset = queryset.filter(priority=priority)
        if date_from:
            queryset = queryset.filter(report_date__gte=date_from)
        if date_to:
            queryset = queryset.filter(report_date__lte=date_to)
        if responsible_employee:
            queryset = queryset.filter(responsible_employee_id=responsible_employee)
        queryset = _apply_operational_scope_filter(
            queryset,
            self.request.user,
            department_field="calendar__department_id",
            process_field="process_id",
            shift_field="shift_id",
        )
        return queryset


class HikitsuguiItemViewSet(ActorMixin, viewsets.ModelViewSet):
    queryset = HikitsuguiItem.objects.select_related("report", "category", "responsible_employee")
    serializer_class = HikitsuguiItemSerializer
    permission_classes = [HikitsuguiPermission]

    def get_queryset(self):
        queryset = super().get_queryset()
        params = self.request.query_params
        report = params.get("report")
        status_param = params.get("status")
        priority = params.get("priority")
        if report:
            queryset = queryset.filter(report_id=report)
        if status_param:
            queryset = queryset.filter(status=status_param)
        if priority:
            queryset = queryset.filter(priority=priority)
        queryset = _apply_operational_scope_filter(
            queryset,
            self.request.user,
            department_field="report__calendar__department_id",
            process_field="report__process_id",
            shift_field="report__shift_id",
        )
        return queryset


class ProductionMonitorSourceViewSet(ActorMixin, viewsets.ModelViewSet):
    queryset = ProductionMonitorSource.objects.select_related("process")
    serializer_class = ProductionMonitorSourceSerializer
    permission_classes = [OperationsCalendarPermission]


class ProductionSnapshotViewSet(ActorMixin, viewsets.ModelViewSet):
    queryset = ProductionSnapshot.objects.select_related("source", "process", "shift").prefetch_related("machine_statuses")
    serializer_class = ProductionSnapshotSerializer
    permission_classes = [OperationsCalendarPermission]

    def get_queryset(self):
        queryset = super().get_queryset()
        params = self.request.query_params
        process = params.get("process")
        area = params.get("area")
        shift = params.get("shift")
        status_param = params.get("status")
        if process:
            queryset = queryset.filter(process_id=process)
        if area:
            queryset = queryset.filter(area=area)
        if shift:
            queryset = queryset.filter(shift_id=shift)
        if status_param:
            queryset = queryset.filter(machine_statuses__status=status_param).distinct()
        return queryset

    @action(detail=False, methods=["get"], url_path="dashboard")
    def dashboard(self, request):
        queryset = self.get_queryset()
        latest = queryset.first()

        if not latest:
            mock_machines = [
                {
                    "id": 0,
                    "snapshot": None,
                    "machine_code": "MCH-01",
                    "equipment_name": "Linha A - Prensa 01",
                    "status": "running",
                    "production_actual": 1220,
                    "production_target": 1300,
                    "difference": -80,
                    "kadouritsu": "93.85",
                    "run_minutes": 430,
                    "stop_minutes": 28,
                    "last_update_at": datetime.now().isoformat(),
                    "alarm_active": False,
                },
                {
                    "id": 0,
                    "snapshot": None,
                    "machine_code": "MCH-02",
                    "equipment_name": "Linha A - Solda 02",
                    "status": "stopped",
                    "production_actual": 860,
                    "production_target": 1300,
                    "difference": -440,
                    "kadouritsu": "66.15",
                    "run_minutes": 305,
                    "stop_minutes": 153,
                    "last_update_at": datetime.now().isoformat(),
                    "alarm_active": True,
                },
                {
                    "id": 0,
                    "snapshot": None,
                    "machine_code": "MCH-03",
                    "equipment_name": "Linha B - Montagem 03",
                    "status": "idle",
                    "production_actual": 0,
                    "production_target": 980,
                    "difference": -980,
                    "kadouritsu": "0.00",
                    "run_minutes": 0,
                    "stop_minutes": 458,
                    "last_update_at": datetime.now().isoformat(),
                    "alarm_active": False,
                },
            ]
            return Response(
                {
                    "snapshot": None,
                    "kpis": {
                        "production_total": 2080,
                        "target_total": 3580,
                        "difference_total": -1500,
                        "average_kadouritsu": 53.33,
                        "running_count": 1,
                        "stopped_count": 1,
                        "idle_count": 1,
                        "error_count": 0,
                        "alarms_active": 1,
                    },
                    "machines": mock_machines,
                    "is_mock": True,
                }
            )

        machines_qs = latest.machine_statuses.all()
        status_param = request.query_params.get("status")
        if status_param:
            machines_qs = machines_qs.filter(status=status_param)

        machines_data = ProductionMachineStatusSerializer(machines_qs, many=True).data
        metrics = getattr(latest, "metrics", None)
        if metrics:
            metrics_data = ProductionMetricsSerializer(metrics).data
            kpis = {
                "production_total": metrics_data["total_actual"],
                "target_total": metrics_data["total_target"],
                "difference_total": (metrics_data["total_actual"] or 0) - (metrics_data["total_target"] or 0),
                "average_kadouritsu": metrics_data["average_kadouritsu"],
                "running_count": metrics_data["running_count"],
                "stopped_count": metrics_data["stopped_count"],
                "idle_count": metrics_data["idle_count"],
                "error_count": metrics_data["error_count"],
                "alarms_active": metrics_data["alarms_active"],
            }
        else:
            production_total = sum((item["production_actual"] or 0) for item in machines_data)
            target_total = sum((item["production_target"] or 0) for item in machines_data)
            kadouritsu_values = [float(item["kadouritsu"]) for item in machines_data if item.get("kadouritsu") is not None]
            kpis = {
                "production_total": production_total,
                "target_total": target_total,
                "difference_total": production_total - target_total,
                "average_kadouritsu": round((sum(kadouritsu_values) / len(kadouritsu_values)) if kadouritsu_values else 0, 2),
                "running_count": sum(1 for item in machines_data if item["status"] == "running"),
                "stopped_count": sum(1 for item in machines_data if item["status"] == "stopped"),
                "idle_count": sum(1 for item in machines_data if item["status"] == "idle"),
                "error_count": sum(1 for item in machines_data if item["status"] == "error"),
                "alarms_active": sum(1 for item in machines_data if item["alarm_active"]),
            }

        return Response(
            {
                "snapshot": ProductionSnapshotSerializer(latest).data,
                "kpis": kpis,
                "machines": machines_data,
                "is_mock": False,
            }
        )


class ProductionMachineStatusViewSet(ActorMixin, viewsets.ReadOnlyModelViewSet):
    queryset = ProductionMachineStatus.objects.select_related("snapshot")
    serializer_class = ProductionMachineStatusSerializer
    permission_classes = [OperationsCalendarPermission]


class ProductionMetricsViewSet(ActorMixin, viewsets.ReadOnlyModelViewSet):
    queryset = ProductionMetrics.objects.select_related("snapshot")
    serializer_class = ProductionMetricsSerializer
    permission_classes = [OperationsCalendarPermission]


class OperationsSettingsViewSet(ActorMixin, viewsets.ModelViewSet):
    queryset = OperationsSettings.objects.all()
    serializer_class = OperationsSettingsSerializer
    permission_classes = [OperationsSettingsPermission]

    def _safe_load_settings(self):
        try:
            return OperationsSettings.load()
        except (OperationalError, ProgrammingError):
            return OperationsSettings(
                weekly_warning_hours=50,
                weekly_critical_hours=60,
                monthly_overtime_warning_hours=45,
                monthly_overtime_critical_hours=60,
                consecutive_absence_warning=2,
                recurrent_late_warning=3,
                enable_kajuuroudou_alerts=True,
                notes="",
            )

    def get_queryset(self):
        settings_obj = self._safe_load_settings()
        if settings_obj.pk:
            return OperationsSettings.objects.filter(pk=settings_obj.pk)
        return OperationsSettings.objects.none()

    @action(detail=False, methods=["get", "patch"], url_path="current")
    def current(self, request):
        settings_obj = self._safe_load_settings()
        if request.method == "GET":
            return Response(self.get_serializer(settings_obj).data)
        if not settings_obj.pk:
            return Response({"detail": "Configuração ainda não disponível no banco."}, status=status.HTTP_503_SERVICE_UNAVAILABLE)
        serializer = self.get_serializer(settings_obj, data=request.data, partial=True)
        serializer.is_valid(raise_exception=True)
        serializer.save(updated_by=self._actor())
        return Response(serializer.data)


class EmployeeAdministrativeNoteViewSet(ActorMixin, viewsets.ModelViewSet):
    queryset = EmployeeAdministrativeNote.objects.select_related("employee", "created_by", "updated_by")
    serializer_class = EmployeeAdministrativeNoteSerializer
    permission_classes = [EmployeeAdminNotePermission]

    def get_queryset(self):
        queryset = super().get_queryset()
        employee_id = self.request.query_params.get("employee")
        if employee_id:
            queryset = queryset.filter(employee__employee_id=employee_id)
        return queryset


class AttendanceDashboardViewSet(viewsets.ViewSet):
    permission_classes = [AttendanceDashboardPermission]

    def _safe_load_settings(self):
        try:
            return OperationsSettings.load()
        except (OperationalError, ProgrammingError):
            return OperationsSettings(
                weekly_warning_hours=50,
                weekly_critical_hours=60,
                monthly_overtime_warning_hours=45,
                monthly_overtime_critical_hours=60,
                consecutive_absence_warning=2,
                recurrent_late_warning=3,
                enable_kajuuroudou_alerts=True,
                notes="",
            )

    def _employee_display_name(self, employee):
        return (
            getattr(employee, "name_en", None)
            or getattr(employee, "internal_name", None)
            or getattr(employee, "name_jp", None)
            or ""
        )

    def _timecard_scope(self, request, queryset):
        relevant_codes = set()
        cell_map = {}
        for cell in queryset.select_related("assignment__employee"):
            employee = getattr(cell.assignment, "employee", None)
            if not employee:
                continue
            normalized_code = _normalize_employee_code(getattr(employee, "employee_cd", "") or getattr(employee, "employee_id", ""))
            if not normalized_code:
                continue
            relevant_codes.add(normalized_code)
            cell_map[(normalized_code, cell.date)] = cell

        if not relevant_codes:
            return {
                "records": [],
                "record_map": {},
                "cell_map": cell_map,
                "relevant_codes": relevant_codes,
                "divergences": [],
                "summary": {
                    "total_records": 0,
                    "matched_records": 0,
                    "unmatched_records": 0,
                    "divergences_count": 0,
                    "missing_punch_count": 0,
                    "worked_on_day_off_count": 0,
                    "late_count": 0,
                    "early_leave_count": 0,
                    "overtime_mismatch_count": 0,
                    "work_minutes_mismatch_count": 0,
                },
            }

        records_qs = AttendanceTimecardRecord.objects.filter(employee_code_normalized__in=relevant_codes)
        month = request.query_params.get("month")
        date_from = request.query_params.get("date_from")
        date_to = request.query_params.get("date_to")
        if date_from:
            records_qs = records_qs.filter(work_date__gte=date_from)
        if date_to:
            records_qs = records_qs.filter(work_date__lte=date_to)
        if not date_from and not date_to and month:
            try:
                year_str, month_str = month.split("-")
                month_year = int(year_str)
                month_num = int(month_str)
                records_qs = records_qs.filter(work_date__year=month_year, work_date__month=month_num)
            except (TypeError, ValueError):
                pass
        elif not date_from and not date_to:
            bounds = queryset.aggregate(min_date=Min("date"), max_date=Max("date"))
            if bounds.get("min_date"):
                records_qs = records_qs.filter(work_date__gte=bounds["min_date"], work_date__lte=bounds["max_date"])

        records = list(records_qs.order_by("work_date", "employee_code_normalized"))
        record_map = {(record.employee_code_normalized, record.work_date): record for record in records}

        divergences = []
        calendar_ids = list(queryset.values_list("calendar_id", flat=True).distinct())
        calendars = MonthlyOperationCalendar.objects.filter(id__in=calendar_ids)
        for calendar in calendars:
            divergences.extend(compare_timecard_to_calendar(calendar, records=records))

        matched_records = sum(1 for record in records if (record.employee_code_normalized, record.work_date) in cell_map)
        summary = {
            "total_records": len(records),
            "matched_records": matched_records,
            "unmatched_records": max(len(records) - matched_records, 0),
            "divergences_count": len(divergences),
            "missing_punch_count": sum(1 for item in divergences if item.get("type") == "missing_timecard"),
            "worked_on_day_off_count": sum(1 for item in divergences if item.get("type") == "worked_on_day_off"),
            "late_count": sum(1 for item in divergences if item.get("type") == "late"),
            "early_leave_count": sum(1 for item in divergences if item.get("type") == "early_leave"),
            "overtime_mismatch_count": sum(1 for item in divergences if item.get("type") == "overtime_mismatch"),
            "work_minutes_mismatch_count": sum(1 for item in divergences if item.get("type") == "work_minutes_mismatch"),
        }
        return {
            "records": records,
            "record_map": record_map,
            "cell_map": cell_map,
            "relevant_codes": relevant_codes,
            "divergences": divergences,
            "summary": summary,
        }

    def _serialize_timecard_divergence(self, item, cell_map=None, record_map=None):
        cell_map = cell_map or {}
        record_map = record_map or {}
        employee_code = _normalize_employee_code(item.get("employee_code"))
        record = record_map.get((employee_code, item.get("date")))
        cell = cell_map.get((employee_code, item.get("date")))
        employee = getattr(getattr(cell, "assignment", None), "employee", None)
        employee_id = getattr(employee, "employee_id", None) if employee else None
        employee_name = (
            item.get("employee_name")
            or (self._employee_display_name(employee) if employee else "")
            or getattr(record, "employee_name", "")
            or ""
        )
        return {
            "employee_id": employee_id,
            "employee_code": employee_code,
            "employee_name": employee_name,
            "date": item.get("date"),
            "type": item.get("type"),
            "severity": item.get("severity") or "warning",
            "expected": item.get("expected"),
            "actual": item.get("actual"),
            "message": item.get("message"),
        }

    def _filtered_cells_queryset(self, request, *, apply_scope=True):
        params = request.query_params
        month = params.get("month")
        date_from = params.get("date_from")
        date_to = params.get("date_to")
        department = params.get("department")
        process = params.get("process")
        shift = params.get("shift")
        rotation_group = params.get("group")

        queryset = CalendarDayCell.objects.select_related(
            "calendar",
            "assignment",
            "assignment__employee",
            "assignment__employee__process",
            "assignment__employee__shift",
            "attendance_status",
            "work_time_code",
            "operational_code",
            "position",
        )

        if month:
            try:
                year_str, month_str = month.split("-")
                queryset = queryset.filter(date__year=int(year_str), date__month=int(month_str))
            except (ValueError, TypeError):
                raise ValueError("month inválido. Use YYYY-MM.")

        if date_from:
            queryset = queryset.filter(date__gte=date_from)
        if date_to:
            queryset = queryset.filter(date__lte=date_to)
        if department:
            queryset = queryset.filter(calendar__department_id=department)
        if process:
            queryset = queryset.filter(calendar__process_id=process)
        if shift:
            queryset = queryset.filter(calendar__shift_id=shift)
        if rotation_group:
            queryset = queryset.filter(assignment__rotation_group=rotation_group)
        if apply_scope:
            queryset = _apply_operational_scope_filter(
                queryset,
                request.user,
                department_field="calendar__department_id",
                process_field="calendar__process_id",
                shift_field="calendar__shift_id",
            )
        return queryset

    def list(self, request):
        settings_obj = self._safe_load_settings()
        try:
            queryset = self._filtered_cells_queryset(request)
        except ValueError as exc:
            return Response({"detail": str(exc)}, status=status.HTTP_400_BAD_REQUEST)

        total_employees = queryset.values("assignment_id").distinct().count()
        total_records = queryset.count()

        present_count = queryset.filter(
            Q(attendance_status__is_working_day=True) & Q(attendance_status__is_absence=False)
        ).count()
        absence_count = queryset.filter(attendance_status__is_absence=True).count()
        off_count = queryset.filter(
            Q(attendance_status__is_working_day=False) & Q(attendance_status__is_absence=False)
        ).count()
        late_count = queryset.filter(
            Q(attendance_status__code__icontains="late") | Q(operational_code__code__icontains="chikoku")
        ).count()
        early_leave_count = queryset.filter(
            Q(attendance_status__code__icontains="early") | Q(operational_code__code__icontains="soutai")
        ).count()
        no_status_count = queryset.filter(attendance_status__isnull=True).count()

        overtime_day_minutes = queryset.aggregate(total=Sum("overtime_minutes"))["total"] or 0

        week_start = datetime.now().date() - timedelta(days=datetime.now().date().weekday())
        month_start = datetime.now().date().replace(day=1)
        overtime_week_minutes = queryset.filter(date__gte=week_start).aggregate(total=Sum("overtime_minutes"))["total"] or 0
        overtime_month_minutes = queryset.filter(date__gte=month_start).aggregate(total=Sum("overtime_minutes"))["total"] or 0

        per_assignment = (
            queryset.values("assignment_id", "assignment__employee__employee_id", "assignment__employee__name_en")
            .annotate(
                absences=Count("id", filter=Q(attendance_status__is_absence=True)),
                lates=Count("id", filter=Q(attendance_status__code__icontains="late") | Q(operational_code__code__icontains="chikoku")),
                no_status=Count("id", filter=Q(attendance_status__isnull=True)),
                overtime_minutes=Sum("overtime_minutes"),
                actual_minutes=Sum("actual_work_minutes"),
            )
        )

        by_assignment = list(per_assignment)
        by_assignment_sorted_absence = sorted(by_assignment, key=lambda item: item["absences"] or 0, reverse=True)
        by_assignment_sorted_late = sorted(by_assignment, key=lambda item: item["lates"] or 0, reverse=True)
        by_assignment_sorted_ot = sorted(by_assignment, key=lambda item: item["overtime_minutes"] or 0, reverse=True)

        absences_consecutive = []
        recurring_lates = []
        consecutive_absence_threshold = int(settings_obj.consecutive_absence_warning or 2)
        recurrent_late_threshold = int(settings_obj.recurrent_late_warning or 3)
        for row in by_assignment:
            aid = row["assignment_id"]
            timeline = list(
                queryset.filter(assignment_id=aid)
                .order_by("date")
                .values("date", "attendance_status__is_absence", "attendance_status__code", "operational_code__code")
            )
            max_streak = 0
            current = 0
            late_hits = 0
            for entry in timeline:
                is_abs = bool(entry["attendance_status__is_absence"])
                if is_abs:
                    current += 1
                    max_streak = max(max_streak, current)
                else:
                    current = 0
                status_code = str(entry.get("attendance_status__code") or "").lower()
                op_code = str(entry.get("operational_code__code") or "").lower()
                if "late" in status_code or "chikoku" in op_code:
                    late_hits += 1
            if max_streak >= consecutive_absence_threshold:
                absences_consecutive.append({**row, "max_absence_streak": max_streak})
            if late_hits >= recurrent_late_threshold:
                recurring_lates.append({**row, "late_hits": late_hits})

        risk_alerts = []
        if settings_obj.enable_kajuuroudou_alerts:
            weekly_warning = float(settings_obj.weekly_warning_hours or 50)
            weekly_critical = float(settings_obj.weekly_critical_hours or 60)
            monthly_warning = float(settings_obj.monthly_overtime_warning_hours or 45)
            monthly_critical = float(settings_obj.monthly_overtime_critical_hours or 60)
            for row in by_assignment:
                actual_hours = ((row["actual_minutes"] or 0) / 60.0)
                overtime_hours = ((row["overtime_minutes"] or 0) / 60.0)
                level = None
                reasons = []
                if actual_hours > weekly_critical:
                    level = "critical"
                    reasons.append(f"Acima do limite semanal crítico ({weekly_critical}h)")
                elif actual_hours > weekly_warning:
                    level = "warning"
                    reasons.append(f"Acima do limite semanal ({weekly_warning}h)")
                if overtime_hours > monthly_critical:
                    level = "critical"
                    reasons.append(f"Acima do limite mensal crítico ({monthly_critical}h extras)")
                elif overtime_hours > monthly_warning:
                    level = level or "warning"
                    reasons.append(f"Acima do limite mensal ({monthly_warning}h extras)")
                if overtime_hours > 80:
                    level = "critical"
                    reasons.append("Próximo de 80h extras")

                if level:
                    risk_alerts.append(
                        {
                            "assignment_id": row["assignment_id"],
                            "employee_id": row["assignment__employee__employee_id"],
                            "employee_name": row["assignment__employee__name_en"],
                            "actual_hours": round(actual_hours, 2),
                            "overtime_hours": round(overtime_hours, 2),
                            "level": level,
                            "reasons": reasons,
                        }
                    )

        timecard_payload = self._timecard_scope(request, queryset)
        timecard_divergences = [
            self._serialize_timecard_divergence(
                item,
                cell_map=timecard_payload["cell_map"],
                record_map=timecard_payload["record_map"],
            )
            for item in timecard_payload["divergences"]
        ]

        return Response(
            {
                "kpis": {
                    "total_scheduled_employees": total_employees,
                    "present": present_count,
                    "absences": absence_count,
                    "lates": late_count,
                    "early_leaves": early_leave_count,
                    "offs": off_count,
                    "overtime_day_hours": round(overtime_day_minutes / 60.0, 2),
                    "overtime_week_hours": round(overtime_week_minutes / 60.0, 2),
                    "overtime_month_hours": round(overtime_month_minutes / 60.0, 2),
                    "risk_people": len(risk_alerts),
                },
                "attendance_summary": {
                    "total_records": total_records,
                    "without_status": no_status_count,
                },
                "timecard_summary": timecard_payload["summary"],
                "timecard_divergences": timecard_divergences[:100],
                "overtime_summary": {
                    "by_employee": [
                        {
                            "assignment_id": row["assignment_id"],
                            "employee_id": row["assignment__employee__employee_id"],
                            "employee_name": row["assignment__employee__name_en"],
                            "overtime_hours": round(((row["overtime_minutes"] or 0) / 60.0), 2),
                        }
                        for row in by_assignment_sorted_ot[:30]
                    ]
                },
                "risk_alerts": risk_alerts[:50],
                "settings": OperationsSettingsSerializer(settings_obj).data,
                "employee_rankings": {
                    "most_absences_month": [
                        {
                            "assignment_id": row["assignment_id"],
                            "employee_id": row["assignment__employee__employee_id"],
                            "employee_name": row["assignment__employee__name_en"],
                            "count": row["absences"] or 0,
                        }
                        for row in by_assignment_sorted_absence[:20]
                    ],
                    "most_lates_month": [
                        {
                            "assignment_id": row["assignment_id"],
                            "employee_id": row["assignment__employee__employee_id"],
                            "employee_name": row["assignment__employee__name_en"],
                            "count": row["lates"] or 0,
                        }
                        for row in by_assignment_sorted_late[:20]
                    ],
                    "consecutive_absences": absences_consecutive[:20],
                    "recurring_lates": recurring_lates[:20],
                    "without_status": [
                        {
                            "assignment_id": row["assignment_id"],
                            "employee_id": row["assignment__employee__employee_id"],
                            "employee_name": row["assignment__employee__name_en"],
                            "count": row["no_status"] or 0,
                        }
                        for row in sorted(by_assignment, key=lambda item: item["no_status"] or 0, reverse=True)[:20]
                    ],
                },
            }
        )


class AttendanceDashboardViewSet(AttendanceDashboardViewSet):
    @action(detail=False, methods=["get"], url_path=r"employees/(?P<employee_id>[^/.]+)")
    def employee_detail(self, request, employee_id=None):
        if get_user_operation_role_code(request.user) == "dashboard_tv" and not request.user.is_superuser:
            return Response({"detail": "Você não tem permissão para executar esta ação."}, status=status.HTTP_403_FORBIDDEN)
        settings_obj = self._safe_load_settings()
        try:
            queryset = self._filtered_cells_queryset(request)
        except ValueError as exc:
            return Response({"detail": str(exc)}, status=status.HTTP_400_BAD_REQUEST)
        queryset = queryset.filter(assignment__employee__employee_id=employee_id)
        first_cell = queryset.first()
        if not first_cell and not get_user_operation_profile(request.user):
            try:
                queryset = self._filtered_cells_queryset(request, apply_scope=False).filter(
                    assignment__employee__employee_id=employee_id
                )
                first_cell = queryset.first()
            except ValueError:
                pass
        if not first_cell:
            queryset = CalendarDayCell.objects.select_related(
                "calendar",
                "assignment",
                "assignment__employee",
                "attendance_status",
                "work_time_code",
                "operational_code",
                "position",
            ).filter(assignment__employee__employee_id=employee_id)
            first_cell = queryset.first()
        if not first_cell:
            return Response({"detail": "Funcionário sem registros no filtro informado."}, status=status.HTTP_404_NOT_FOUND)

        employee = first_cell.assignment.employee
        profile = {
            "employee_pk": employee.pk,
            "employee_id": employee.employee_id,
            "name": employee.name_en or employee.internal_name or employee.name_jp or "",
            "process": getattr(getattr(employee, "process", None), "code", None),
            "shift": getattr(getattr(employee, "shift", None), "code", None),
            "group": first_cell.assignment.rotation_group or "",
        }

        absences = queryset.filter(attendance_status__is_absence=True).count()
        lates = queryset.filter(Q(attendance_status__code__icontains="late") | Q(operational_code__code__icontains="chikoku")).count()
        early_leaves = queryset.filter(Q(attendance_status__code__icontains="early") | Q(operational_code__code__icontains="soutai")).count()
        offs = queryset.filter(Q(attendance_status__is_working_day=False) & Q(attendance_status__is_absence=False)).count()

        daily_rows = list(
            queryset.order_by("date").values(
                "date",
                "attendance_status__code",
                "attendance_status__label_pt",
                "operational_code__code",
                "work_time_code__code",
                "overtime_minutes",
                "actual_work_minutes",
                "memo",
                "time_note",
                "raw_value",
            )
        )
        weekly_minutes = sum(item["actual_work_minutes"] or 0 for item in daily_rows)
        monthly_ot_minutes = sum(item["overtime_minutes"] or 0 for item in daily_rows)

        timecard_payload = self._timecard_scope(request, queryset)
        timecard_divergence_map = {
            (item["employee_code"], item["date"]): item for item in timecard_payload["divergences"]
        }
        timecard_records = []
        for record in timecard_payload["records"]:
            divergence = timecard_divergence_map.get((record.employee_code_normalized, record.work_date))
            timecard_records.append(
                {
                    "date": record.work_date,
                    "clock_in": record.clock_in.strftime("%H:%M") if record.clock_in else "",
                    "clock_out": record.clock_out.strftime("%H:%M") if record.clock_out else "",
                    "total_work_hours": round((record.total_work_minutes or 0) / 60.0, 2),
                    "scheduled_work_hours": round((record.scheduled_work_minutes or 0) / 60.0, 2),
                    "overtime_hours": round((record.overtime_minutes or 0) / 60.0, 2),
                    "late_minutes": int(record.late_minutes or 0),
                    "early_leave_minutes": int(record.early_leave_minutes or 0),
                    "memo": record.memo or record.work_type_name or record.shift_name or "",
                    "divergence_type": divergence["type"] if divergence else None,
                    "divergence_severity": divergence["severity"] if divergence else None,
                    "divergence_message": divergence["message"] if divergence else None,
                    "divergence_expected": divergence["expected"] if divergence else None,
                    "divergence_actual": divergence["actual"] if divergence else None,
                }
            )

        risk_alerts = []
        weekly_warning = float(settings_obj.weekly_warning_hours or 50)
        weekly_critical = float(settings_obj.weekly_critical_hours or 60)
        monthly_warning = float(settings_obj.monthly_overtime_warning_hours or 45)
        monthly_critical = float(settings_obj.monthly_overtime_critical_hours or 60)
        weekly_hours = weekly_minutes / 60.0
        monthly_ot_hours = monthly_ot_minutes / 60.0
        if settings_obj.enable_kajuuroudou_alerts:
            if weekly_hours > weekly_critical:
                risk_alerts.append(f"Acima do limite semanal crítico ({weekly_critical}h)")
            elif weekly_hours > weekly_warning:
                risk_alerts.append(f"Acima do limite semanal ({weekly_warning}h)")
            if monthly_ot_hours > monthly_critical:
                risk_alerts.append(f"Acima do limite mensal crítico ({monthly_critical}h extras)")
            elif monthly_ot_hours > monthly_warning:
                risk_alerts.append(f"Acima do limite mensal ({monthly_warning}h extras)")

        return Response(
            {
                "employee": profile,
                "summary": {
                    "absences": absences,
                    "lates": lates,
                    "early_leaves": early_leaves,
                    "offs": offs,
                    "weekly_overtime_hours": round(monthly_ot_minutes / 60.0, 2),
                    "monthly_overtime_hours": round(monthly_ot_minutes / 60.0, 2),
                    "weekly_worked_hours": round(weekly_hours, 2),
                },
                "risk_alerts": risk_alerts,
                "timecard_records": timecard_records,
                "administrative_notes": EmployeeAdministrativeNoteSerializer(
                    EmployeeAdministrativeNote.objects.filter(employee=employee).select_related("created_by")[:20],
                    many=True,
                ).data,
                "daily_history": [
                    {
                        "date": row["date"],
                        "status": row["attendance_status__label_pt"] or row["attendance_status__code"] or "-",
                        "status_code": row["attendance_status__code"],
                        "operational_code": row["operational_code__code"],
                        "work_time_code": row["work_time_code__code"],
                        "actual_work_hours": round((row["actual_work_minutes"] or 0) / 60.0, 2),
                        "overtime_hours": round((row["overtime_minutes"] or 0) / 60.0, 2),
                        "note": row["memo"] or row["time_note"] or row["raw_value"] or "",
                    }
                    for row in daily_rows
                ],
            }
        )


class OperationsMePermissionViewSet(viewsets.ViewSet):
    permission_classes = [IsAuthenticated]

    def list(self, request):
        return Response(get_user_operation_permissions_payload(request.user))


class OperationsAccessManagementViewSet(ActorMixin, viewsets.ViewSet):
    permission_classes = [OperationsRBACManagementPermission]
    UserModel = get_user_model()

    def _serialize_scope(self, scope):
        return UserOperationScopeSerializer(scope).data

    def _build_user_row(self, user):
        profile = getattr(user, "operation_profile", None)
        scopes = []
        role_code = None
        additional_roles = []
        updated_at = None
        profile_id = None
        if profile and profile.is_active:
            profile_id = profile.id
            role_code = getattr(profile.role, "code", None)
            additional_roles = list(profile.additional_roles.filter(is_active=True).values_list("code", flat=True))
            scopes = [self._serialize_scope(s) for s in profile.scopes.filter(is_active=True).select_related("role", "department", "process", "shift")]
            updated_at = profile.updated_at

        full_name = (user.get_full_name() or "").strip()
        return {
            "user_id": user.id,
            "username": user.username,
            "full_name": full_name or user.username,
            "is_active": user.is_active,
            "operation_profile_id": profile_id,
            "role": role_code,
            "additional_roles": additional_roles,
            "scopes": scopes,
            "updated_at": updated_at,
        }

    def _get_or_create_profile(self, user):
        profile = getattr(user, "operation_profile", None)
        if profile:
            return profile

        fallback_role = OperationRole.objects.filter(code="viewer", is_active=True).first() or OperationRole.objects.filter(is_active=True).first()
        if not fallback_role:
            raise ValueError("Nenhum OperationRole ativo encontrado.")
        return UserOperationProfile.objects.create(user=user, role=fallback_role, created_by=self._actor(), updated_by=self._actor())

    def _audit(self, *, target_user, action, before_data, after_data):
        OperationAccessAuditLog.objects.create(
            target_user=target_user,
            action=action,
            payload_before=before_data or {},
            payload_after=after_data or {},
            created_by=self._actor(),
            updated_by=self._actor(),
        )

    @action(detail=False, methods=["get"], url_path="users")
    def users(self, request):
        role_filter = (request.query_params.get("role") or "").strip()
        department_filter = request.query_params.get("department")
        process_filter = request.query_params.get("process")
        active_filter = request.query_params.get("active")

        queryset = self.UserModel.objects.all().order_by("username")
        if active_filter in {"true", "false"}:
            queryset = queryset.filter(is_active=(active_filter == "true"))

        queryset = queryset.select_related("operation_profile", "operation_profile__role").prefetch_related(
            "operation_profile__additional_roles",
            "operation_profile__scopes__role",
            "operation_profile__scopes__department",
            "operation_profile__scopes__process",
            "operation_profile__scopes__shift",
        )

        rows = []
        for user in queryset:
            row = self._build_user_row(user)
            if role_filter and row["role"] != role_filter and role_filter not in row["additional_roles"]:
                continue
            if department_filter and not any(str(s.get("department")) == str(department_filter) for s in row["scopes"]):
                continue
            if process_filter and not any(str(s.get("process")) == str(process_filter) for s in row["scopes"]):
                continue
            rows.append(row)

        serializer = OperationAccessUserListSerializer(rows, many=True)
        return Response(serializer.data)

    @action(detail=False, methods=["get"], url_path="meta")
    def meta(self, request):
        return Response(
            {
                "roles": OperationRoleSerializer(OperationRole.objects.filter(is_active=True), many=True).data,
                "departments": [{"id": d.id, "code": d.code, "label": d.label_pt or d.label_jp} for d in Department.objects.all().order_by("code")],
                "processes": [{"id": p.id, "code": p.code, "label": p.label_pt or p.label_jp} for p in Process.objects.all().order_by("code")],
                "shifts": [{"id": s.id, "code": s.code, "label": s.label_pt or s.label_jp} for s in Shift.objects.all().order_by("code")],
            }
        )

    @action(detail=False, methods=["get", "patch"], url_path=r"users/(?P<user_id>\d+)/profile")
    def user_profile(self, request, user_id=None):
        user = get_object_or_404(self.UserModel, pk=user_id)
        profile = self._get_or_create_profile(user)

        if request.method == "GET":
            return Response(UserOperationProfileSerializer(profile).data)

        before_data = UserOperationProfileSerializer(profile).data
        serializer = UserOperationProfileSerializer(profile, data=request.data, partial=True)
        serializer.is_valid(raise_exception=True)
        serializer.save(updated_by=self._actor())
        after_data = UserOperationProfileSerializer(profile).data
        self._audit(
            target_user=user,
            action=OperationAccessAuditLog.Action.PROFILE_UPDATED,
            before_data=before_data,
            after_data=after_data,
        )
        return Response(after_data)

    @action(detail=False, methods=["put"], url_path=r"users/(?P<user_id>\d+)/scopes")
    def replace_scopes(self, request, user_id=None):
        user = get_object_or_404(self.UserModel, pk=user_id)
        profile = self._get_or_create_profile(user)
        scope_items = request.data if isinstance(request.data, list) else request.data.get("scopes", [])
        if not isinstance(scope_items, list):
            return Response({"detail": "Formato inválido para scopes."}, status=status.HTTP_400_BAD_REQUEST)

        before_scopes = UserOperationScopeSerializer(
            profile.scopes.filter(is_active=True).select_related("role", "department", "process", "shift"),
            many=True,
        ).data

        with transaction.atomic():
            profile.scopes.filter(is_active=True).update(is_active=False, updated_by=self._actor())
            created = []
            for item in scope_items:
                serializer = UserOperationScopeSerializer(data=item)
                serializer.is_valid(raise_exception=True)
                created.append(
                    serializer.save(profile=profile, created_by=self._actor(), updated_by=self._actor())
                )

        after_scopes = UserOperationScopeSerializer(created, many=True).data
        self._audit(
            target_user=user,
            action=OperationAccessAuditLog.Action.SCOPES_REPLACED,
            before_data={"scopes": before_scopes},
            after_data={"scopes": after_scopes},
        )
        return Response({"scopes": after_scopes})

    @action(detail=False, methods=["get"], url_path="audit")
    def audit(self, request):
        user_id = request.query_params.get("user")
        queryset = OperationAccessAuditLog.objects.select_related("created_by", "target_user")
        if user_id:
            queryset = queryset.filter(target_user_id=user_id)
        queryset = queryset[:200]
        return Response(OperationAccessAuditLogSerializer(queryset, many=True).data)


class MonthlyOperationCalendarViewSet(ActorMixin, viewsets.ModelViewSet):
    queryset = MonthlyOperationCalendar.objects.select_related("department", "process", "shift")
    serializer_class = MonthlyOperationCalendarSerializer
    permission_classes = [ScheduleWritePermission]

    def get_queryset(self):
        queryset = super().get_queryset()
        return _apply_operational_scope_filter(
            queryset,
            self.request.user,
            department_field="department_id",
            process_field="process_id",
            shift_field="shift_id",
        )

    def _ordered_assignments_queryset(self, calendar):
        queryset = calendar.assignments.select_related(
            "employee",
            "employee__process",
            "employee__billing_rate",
        )
        return sorted(queryset, key=get_assignment_sort_key)

    @action(detail=True, methods=["get", "post"])
    def assignments(self, request, pk=None):
        calendar = self.get_object()

        if request.method == "GET":
            queryset = self._ordered_assignments_queryset(calendar)
            serializer = CalendarEmployeeAssignmentSerializer(queryset, many=True)
            return Response(serializer.data)

        serializer = CalendarEmployeeAssignmentSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        user = self._actor()
        assignment = serializer.save(calendar=calendar, created_by=user, updated_by=user)
        return Response(CalendarEmployeeAssignmentSerializer(assignment).data, status=status.HTTP_201_CREATED)

    @action(detail=True, methods=["patch"], url_path=r"assignments/(?P<assignment_id>\d+)")
    def assignment_detail(self, request, pk=None, assignment_id=None):
        calendar = self.get_object()
        assignment = get_object_or_404(CalendarEmployeeAssignment, pk=assignment_id, calendar=calendar)
        serializer = CalendarEmployeeAssignmentSerializer(assignment, data=request.data, partial=True)
        serializer.is_valid(raise_exception=True)
        serializer.save(updated_by=self._actor())
        return Response(CalendarEmployeeAssignmentSerializer(assignment).data)

    @action(detail=True, methods=["get", "post"])
    def cells(self, request, pk=None):
        calendar = self.get_object()
        history_source = self._normalize_history_source(request.data.get("history_source"))

        if request.method == "GET":
            queryset = self._calendar_cells(calendar)
            serializer = CalendarDayCellSerializer(queryset, many=True)
            return Response(serializer.data)

        payload = request.data.copy()
        payload.pop("history_source", None)
        serializer = CalendarDayCellSerializer(data=payload)
        serializer.is_valid(raise_exception=True)
        assignment = serializer.validated_data["assignment"]
        if assignment.calendar_id != calendar.id:
            return Response(
                {"assignment": "Assignment does not belong to this calendar."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        user = self._actor()
        old_payload = None
        cell = serializer.save(calendar=calendar, created_by=user, updated_by=user)
        calculate_cell_work_minutes(cell)
        self._log_cell_history(
            calendar=calendar,
            assignment=cell.assignment,
            cell_date=cell.date,
            source=history_source,
            old_value=old_payload,
            new_value=self._snapshot_cell(cell),
            metadata={"origin": "cells_create"},
        )
        return Response(CalendarDayCellSerializer(cell).data, status=status.HTTP_201_CREATED)

    @action(detail=True, methods=["patch"], url_path=r"cells/(?P<cell_id>\d+)")
    def cell_detail(self, request, pk=None, cell_id=None):
        calendar = self.get_object()
        cell = get_object_or_404(CalendarDayCell, pk=cell_id, calendar=calendar)
        old_payload = self._snapshot_cell(cell)
        history_source = self._normalize_history_source(request.data.get("history_source"))
        payload = request.data.copy()
        payload.pop("history_source", None)
        serializer = CalendarDayCellSerializer(cell, data=payload, partial=True)
        serializer.is_valid(raise_exception=True)

        assignment = serializer.validated_data.get("assignment")
        if assignment and assignment.calendar_id != calendar.id:
            return Response(
                {"assignment": "Assignment does not belong to this calendar."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        serializer.save(updated_by=self._actor())
        calculate_cell_work_minutes(cell)
        self._log_cell_history(
            calendar=calendar,
            assignment=cell.assignment,
            cell_date=cell.date,
            source=history_source,
            old_value=old_payload,
            new_value=self._snapshot_cell(cell),
            metadata={"origin": "cells_patch", "cell_id": cell.id},
        )
        return Response(serializer.data)

    @action(detail=True, methods=["post"], url_path="cells/paste")
    def paste_cells(self, request, pk=None):
        calendar = self.get_object()
        start_assignment_id = request.data.get("start_assignment")
        start_date_raw = request.data.get("start_date")
        tsv = request.data.get("tsv", "")

        if not start_assignment_id or not start_date_raw or not isinstance(tsv, str):
            return Response(
                {"detail": "start_assignment, start_date and tsv are required."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            start_date = datetime.strptime(start_date_raw, "%Y-%m-%d").date()
        except ValueError:
            return Response({"start_date": "Invalid date."}, status=status.HTTP_400_BAD_REQUEST)

        if start_date.year != calendar.year or start_date.month != calendar.month:
            return Response(
                {"start_date": "Start date must be inside the calendar month."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        assignments = list(calendar.assignments.order_by("display_order", "employee_id", "id"))
        start_index = next(
            (index for index, assignment in enumerate(assignments) if assignment.id == int(start_assignment_id)),
            None,
        )
        if start_index is None:
            return Response(
                {"start_assignment": "Assignment does not belong to this calendar."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        result = self._paste_tsv(
            calendar,
            assignments[start_index:],
            start_date,
            tsv,
            source=self._normalize_history_source(request.data.get("history_source")) or OperationCalendarHistory.Source.PASTE,
        )
        return Response(result, status=status.HTTP_200_OK)

    @action(detail=True, methods=["get"], url_path="assignment-totals")
    def assignment_totals(self, request, pk=None):
        calendar = self.get_object()
        return Response(recalculate_calendar_totals(calendar))

    @action(detail=True, methods=["get"], url_path="history")
    def history(self, request, pk=None):
        calendar = self.get_object()
        queryset = calendar.history_entries.select_related("assignment__employee", "updated_by")
        assignment_id = request.query_params.get("assignment")
        cell_date = request.query_params.get("date")
        source = request.query_params.get("source")
        user_id = request.query_params.get("user")
        if assignment_id:
            queryset = queryset.filter(assignment_id=assignment_id)
        if cell_date:
            queryset = queryset.filter(cell_date=cell_date)
        if source:
            queryset = queryset.filter(source=source)
        if user_id:
            queryset = queryset.filter(updated_by_id=user_id)
        serializer = OperationCalendarHistorySerializer(queryset[:200], many=True)
        return Response(serializer.data)

    @action(detail=True, methods=["get"], url_path="export-excel")
    def export_excel(self, request, pk=None):
        calendar = self.get_object()
        assignments = list(calendar.assignments.select_related("employee").order_by("display_order", "employee_id", "id"))
        cells = list(
            calendar.day_cells.select_related(
                "assignment",
                "position",
                "attendance_status",
                "work_time_code",
                "operational_code",
            ).order_by("assignment__display_order", "date")
        )
        totals = {row["assignment"]: row for row in recalculate_calendar_totals(calendar)}
        cell_map = {(c.assignment_id, c.date): c for c in cells}
        days = self._month_days(calendar.year, calendar.month)

        wb = Workbook()
        ws = wb.active
        ws.title = "Escala"

        title = f"Escala Operacional {calendar.year}-{str(calendar.month).zfill(2)}"
        scope = f"Depto: {getattr(calendar.department, 'code', calendar.department_id)} | Proc: {getattr(calendar.process, 'code', '-') if calendar.process_id else '-'} | Turno: {getattr(calendar.shift, 'code', '-') if calendar.shift_id else '-'}"
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8 + len(days))
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=8 + len(days))
        ws.cell(row=1, column=1, value=title).font = Font(bold=True, size=14)
        ws.cell(row=2, column=1, value=scope).font = Font(bold=False, size=10)

        header_row = 4
        headers = ["SCD", "Nome", "和名", "Código", "Categoria", "Horas Normais", "Horas Extras", "Acumulado"]
        for idx, label in enumerate(headers, start=1):
            cell = ws.cell(row=header_row, column=idx, value=label)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.fill = PatternFill(fill_type="solid", fgColor="D9E8EF")

        for day_idx, day in enumerate(days, start=9):
            weekday = day.strftime("%a")
            label = f"{day.day}\n{weekday}"
            cell = ws.cell(row=header_row, column=day_idx, value=label)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            fill_color = "FFE9EF" if day.weekday() == 6 else ("EEF2FF" if day.weekday() == 5 else "D9E8EF")
            cell.fill = PatternFill(fill_type="solid", fgColor=fill_color)

        start_data_row = 5
        for row_offset, assignment in enumerate(assignments):
            row_no = start_data_row + row_offset
            employee = assignment.employee
            total = totals.get(assignment.id, {})
            ws.cell(row=row_no, column=1, value=assignment.display_order)
            ws.cell(row=row_no, column=2, value=getattr(employee, "name_en", "") or getattr(employee, "internal_name", "") or getattr(employee, "name_jp", ""))
            ws.cell(row=row_no, column=3, value=getattr(employee, "name_jp", ""))
            ws.cell(row=row_no, column=4, value=getattr(employee, "employee_cd", "") or getattr(employee, "employee_id", ""))
            ws.cell(row=row_no, column=5, value=assignment.operational_category)
            ws.cell(row=row_no, column=6, value=total.get("scheduled_regular_formatted", "0:00"))
            ws.cell(row=row_no, column=7, value=total.get("actual_overtime_formatted", "0:00"))
            ws.cell(row=row_no, column=8, value=total.get("overload_formatted", "0:00"))

            for day_idx, day in enumerate(days, start=9):
                source_cell = cell_map.get((assignment.id, day))
                excel_cell = ws.cell(row=row_no, column=day_idx, value=self._cell_export_text(source_cell))
                excel_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                fill = self._cell_export_fill(source_cell)
                if fill:
                    excel_cell.fill = PatternFill(fill_type="solid", fgColor=fill)

        # Widths
        ws.column_dimensions["A"].width = 6
        ws.column_dimensions["B"].width = 22
        ws.column_dimensions["C"].width = 18
        ws.column_dimensions["D"].width = 12
        ws.column_dimensions["E"].width = 12
        ws.column_dimensions["F"].width = 12
        ws.column_dimensions["G"].width = 12
        ws.column_dimensions["H"].width = 12
        for idx in range(9, 9 + len(days)):
            ws.column_dimensions[get_column_letter(idx)].width = 12

        filename_scope = (getattr(calendar.shift, "code", "") or getattr(calendar.process, "code", "") or "scope").replace(" ", "_")
        filename = f"escala_{calendar.year}_{str(calendar.month).zfill(2)}_{filename_scope}.xlsx"
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        wb.save(response)
        return response

    @action(detail=True, methods=["post"], url_path="save-template")
    def save_template(self, request, pk=None):
        calendar = self.get_object()
        name = str(request.data.get("name") or "").strip()
        if not name:
            return Response({"name": "This field is required."}, status=status.HTTP_400_BAD_REQUEST)

        description = str(request.data.get("description") or "").strip()
        scope_from_calendar = self._parse_bool(request.data.get("scope_from_calendar", True))
        include_base_cells = self._parse_bool(request.data.get("include_base_cells", True))

        with transaction.atomic():
            template = OperationCalendarTemplate.objects.create(
                name=name,
                description=description,
                department=calendar.department if scope_from_calendar else None,
                process=calendar.process if scope_from_calendar else None,
                shift=calendar.shift if scope_from_calendar else None,
                created_by=self._actor(),
                updated_by=self._actor(),
            )
            assignment_map = self._clone_assignments_to_template(calendar, template)
            created_cells = self._clone_cells_to_template_conservative(
                calendar=calendar,
                template=template,
                assignment_map=assignment_map,
                include_base_cells=include_base_cells,
            )
        self._log_operation_history(
            calendar=calendar,
            source=OperationCalendarHistory.Source.TEMPLATE,
            metadata={
                "event": "save_template",
                "template_id": template.id,
                "created_assignments": len(assignment_map),
                "created_cells": created_cells,
            },
        )

        return Response(
            {
                "detail": "Template salvo com sucesso.",
                "template_id": template.id,
                "template_name": template.name,
                "created_assignments": len(assignment_map),
                "created_cells": created_cells,
                "conservative_rules": "Absências, notas e exceções manuais não são replicadas.",
            },
            status=status.HTTP_201_CREATED,
        )

    @action(detail=True, methods=["post"], url_path="apply-template")
    def apply_template(self, request, pk=None):
        calendar = self.get_object()
        template_id = request.data.get("template_id")
        overwrite = self._parse_bool(request.data.get("overwrite", False))
        if not template_id:
            return Response({"template_id": "This field is required."}, status=status.HTTP_400_BAD_REQUEST)

        template = OperationCalendarTemplate.objects.filter(pk=template_id, is_active=True).first()
        if not template:
            return Response({"detail": "Template not found."}, status=status.HTTP_404_NOT_FOUND)

        target_assignment_count = calendar.assignments.count()
        target_cell_count = calendar.day_cells.count()
        target_has_data = target_assignment_count > 0 or target_cell_count > 0
        if target_has_data and not overwrite:
            return Response(
                {
                    "detail": "O calendário de destino já possui dados.",
                    "target_assignment_count": target_assignment_count,
                    "target_cell_count": target_cell_count,
                    "requires_confirmation": True,
                },
                status=status.HTTP_409_CONFLICT,
            )

        with transaction.atomic():
            if target_has_data and overwrite:
                calendar.position_requirements.all().delete()
                calendar.day_cells.all().delete()
                calendar.assignments.all().delete()
            assignment_map = self._clone_template_assignments_to_calendar(template, calendar)
            created_cells = self._clone_template_cells_to_calendar(template, calendar, assignment_map)
        self._log_operation_history(
            calendar=calendar,
            source=OperationCalendarHistory.Source.TEMPLATE,
            metadata={
                "event": "apply_template",
                "template_id": template.id,
                "created_assignments": len(assignment_map),
                "created_cells": created_cells,
                "overwrite": overwrite,
            },
        )

        return Response(
            {
                "detail": "Template aplicado com sucesso.",
                "template_id": template.id,
                "template_name": template.name,
                "target_calendar_id": calendar.id,
                "created_assignments": len(assignment_map),
                "created_cells": created_cells,
            },
            status=status.HTTP_200_OK,
        )

    @action(detail=True, methods=["post"], url_path="duplicate-from-previous")
    def duplicate_from_previous(self, request, pk=None):
        target = self.get_object()
        overwrite = self._parse_bool(request.data.get("overwrite", False))
        copy_base_cells = self._parse_bool(request.data.get("copy_base_cells", True))

        prev_year, prev_month = self._previous_period(target.year, target.month)
        source = MonthlyOperationCalendar.objects.filter(
            department=target.department,
            process=target.process,
            shift=target.shift,
            year=prev_year,
            month=prev_month,
            is_active=True,
        ).first()
        if not source:
            return Response(
                {"detail": "Nenhum calendário do mês anterior encontrado para o mesmo escopo."},
                status=status.HTTP_404_NOT_FOUND,
            )

        target_assignment_count = target.assignments.count()
        target_cell_count = target.day_cells.count()
        target_has_data = target_assignment_count > 0 or target_cell_count > 0
        if target_has_data and not overwrite:
            return Response(
                {
                    "detail": "O calendário de destino já possui dados.",
                    "target_assignment_count": target_assignment_count,
                    "target_cell_count": target_cell_count,
                    "requires_confirmation": True,
                },
                status=status.HTTP_409_CONFLICT,
            )

        with transaction.atomic():
            if target_has_data and overwrite:
                target.position_requirements.all().delete()
                target.day_cells.all().delete()
                target.assignments.all().delete()

            assignment_map = self._clone_assignments(source, target)
            created_cells = self._clone_cells_conservative(
                source=source,
                target=target,
                assignment_map=assignment_map,
                copy_base_cells=copy_base_cells,
            )
        self._log_operation_history(
            calendar=target,
            source=OperationCalendarHistory.Source.MONTH_DUPLICATION,
            metadata={
                "source_calendar_id": source.id,
                "target_calendar_id": target.id,
                "created_assignments": len(assignment_map),
                "created_cells": created_cells,
            },
        )

        return Response(
            {
                "detail": "Duplicação concluída.",
                "source_calendar_id": source.id,
                "target_calendar_id": target.id,
                "created_assignments": len(assignment_map),
                "created_cells": created_cells,
                "copy_base_cells": copy_base_cells,
                "conservative_rules": "Absências, notas e exceções manuais não são replicadas.",
            },
            status=status.HTTP_200_OK,
        )

    @action(detail=True, methods=["post"], url_path="generate-next-month")
    def generate_next_month(self, request, pk=None):
        source = self.get_object()
        copy_assignments = self._parse_bool(request.data.get("copy_assignments", True))
        copy_base_cells = self._parse_bool(request.data.get("copy_base_cells", False))
        overwrite_existing = self._parse_bool(request.data.get("overwrite_existing", False))

        next_year, next_month = self._next_period(source.year, source.month)
        target = MonthlyOperationCalendar.objects.filter(
            department=source.department,
            process=source.process,
            shift=source.shift,
            year=next_year,
            month=next_month,
            is_active=True,
        ).first()

        created_calendar = False
        if not target:
            target = MonthlyOperationCalendar.objects.create(
                department=source.department,
                process=source.process,
                shift=source.shift,
                year=next_year,
                month=next_month,
                title=f"{next_year}-{str(next_month).zfill(2)}",
                status=MonthlyOperationCalendar.Status.DRAFT,
                notes="",
                created_by=self._actor(),
                updated_by=self._actor(),
            )
            created_calendar = True

        target_assignment_count = target.assignments.count()
        target_cell_count = target.day_cells.count()
        target_has_data = target_assignment_count > 0 or target_cell_count > 0
        if target_has_data and not overwrite_existing:
            return Response(
                {
                    "detail": "O calendário do próximo mês já existe e possui dados.",
                    "target_calendar_id": target.id,
                    "target_assignment_count": target_assignment_count,
                    "target_cell_count": target_cell_count,
                    "requires_confirmation": True,
                },
                status=status.HTTP_409_CONFLICT,
            )

        created_assignments = 0
        created_cells = 0
        with transaction.atomic():
            if target_has_data and overwrite_existing:
                target.position_requirements.all().delete()
                target.day_cells.all().delete()
                target.assignments.all().delete()

            assignment_map = {}
            if copy_assignments:
                assignment_map = self._clone_assignments(source, target)
                created_assignments = len(assignment_map)
                created_cells = self._clone_cells_conservative(
                    source=source,
                    target=target,
                    assignment_map=assignment_map,
                    copy_base_cells=copy_base_cells,
                )
        self._log_operation_history(
            calendar=target,
            source=OperationCalendarHistory.Source.NEXT_MONTH_GENERATION,
            metadata={
                "source_calendar_id": source.id,
                "target_calendar_id": target.id,
                "created_assignments": created_assignments,
                "created_cells": created_cells,
            },
        )

        return Response(
            {
                "detail": "Próximo mês gerado com sucesso.",
                "source_calendar_id": source.id,
                "target_calendar_id": target.id,
                "target_year": target.year,
                "target_month": target.month,
                "created_calendar": created_calendar,
                "created_assignments": created_assignments,
                "created_cells": created_cells,
                "copy_assignments": copy_assignments,
                "copy_base_cells": copy_base_cells,
                "conservative_rules": "Absências, notas e exceções manuais não são replicadas.",
            },
            status=status.HTTP_200_OK,
        )

    @action(detail=True, methods=["post"], url_path="generate-schedule")
    def generate_schedule(self, request, pk=None):
        calendar = self.get_object()
        overwrite = self._parse_bool(request.data.get("overwrite", False))
        anchor_date_raw = request.data.get("default_4x2_anchor_date")

        try:
            result = generate_calendar_schedule(
                calendar,
                user=self._actor(),
                overwrite=overwrite,
                default_4x2_anchor_date=anchor_date_raw,
            )
        except ValueError:
            return Response(
                {"default_4x2_anchor_date": "Invalid date."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        self._log_operation_history(
            calendar=calendar,
            source=OperationCalendarHistory.Source.PATTERN_4X2,
            metadata={"result": result},
        )
        return Response(result, status=status.HTTP_200_OK)

    @action(detail=True, methods=["post"], url_path="import-employees")
    def import_employees(self, request, pk=None):
        calendar = self.get_object()
        import_all = self._parse_bool(request.data.get("import_all", False))
        employee_ids = request.data.get("employee_ids") or []

        if not import_all and not employee_ids:
            return Response(
                {"detail": "import_all or employee_ids is required."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        if employee_ids and not isinstance(employee_ids, list):
            return Response({"employee_ids": "Expected a list."}, status=status.HTTP_400_BAD_REQUEST)

        result = import_calendar_employees(
            calendar,
            user=self._actor(),
            import_all=import_all,
            employee_ids=employee_ids,
        )
        return Response(result, status=status.HTTP_200_OK)

    @action(detail=True, methods=["get"], url_path="import-employees-preview")
    def import_employees_preview(self, request, pk=None):
        calendar = self.get_object()
        employee_ids = request.query_params.getlist("employee_ids")
        import_all = self._parse_bool(request.query_params.get("import_all", "true"))
        result = preview_calendar_employee_candidates(
            calendar=calendar,
            import_all=import_all,
            employee_ids=employee_ids,
        )
        return Response(result, status=status.HTTP_200_OK)

    @action(detail=True, methods=["post"], url_path="sync-assignments")
    def sync_assignments(self, request, pk=None):
        calendar = self.get_object()
        result = sync_calendar_assignments_from_master(calendar, user=self._actor())
        self._log_operation_history(
            calendar=calendar,
            source=OperationCalendarHistory.Source.SYSTEM,
            metadata={"event": "sync_assignments", "result": result},
        )
        return Response(result, status=status.HTTP_200_OK)

    def _parse_bool(self, value):
        if isinstance(value, bool):
            return value
        return str(value or "").strip().casefold() in {"1", "true", "yes", "sim"}

    def _month_days(self, year, month):
        from calendar import monthrange
        from datetime import date

        total = monthrange(year, month)[1]
        return [date(year, month, day) for day in range(1, total + 1)]

    def _cell_export_text(self, cell):
        if not cell:
            return ""

        status = getattr(cell, "attendance_status", None)
        if status and (status.is_absence or not status.is_working_day):
            return status.code or status.label_jp or status.label_pt or ""

        position = getattr(cell, "position", None)
        if position:
            floor = getattr(position, "building_floor", None)
            floor_text = getattr(floor, "code", "") or getattr(floor, "label_jp", "") or getattr(floor, "label_pt", "")
            pos_text = getattr(position, "code", "") or getattr(position, "name_jp", "") or getattr(position, "name_pt", "")
            if pos_text and floor_text and floor_text not in pos_text:
                return f"{pos_text} / {floor_text}"
            if pos_text:
                return pos_text

        if getattr(cell, "operational_code", None):
            op = cell.operational_code
            return op.code or op.label_jp or op.label_pt or ""
        if getattr(cell, "work_time_code", None):
            wt = cell.work_time_code
            return wt.code or wt.label_jp or wt.label_pt or ""
        return (cell.raw_value or "").strip()

    def _cell_export_fill(self, cell):
        if not cell:
            return None
        status = getattr(cell, "attendance_status", None)
        op = getattr(cell, "operational_code", None)

        if status:
            if status.is_absence:
                return "FEE2E2"
            if not status.is_working_day:
                return "F3F4F6"
        if op and getattr(op, "category", ""):
            category = str(op.category).lower()
            if "alert" in category or "warn" in category:
                return "FFF3C8"
            if "special" in category or "exception" in category:
                return "EDE9FE"
        return "EDF9EE" if status and status.is_working_day else None

    def _previous_period(self, year, month):
        if month == 1:
            return year - 1, 12
        return year, month - 1

    def _next_period(self, year, month):
        if month == 12:
            return year + 1, 1
        return year, month + 1

    def _clone_assignments(self, source, target):
        user = self._actor()
        assignment_map = {}
        source_assignments = source.assignments.order_by("display_order", "employee_id", "id")
        for source_assignment in source_assignments:
            cloned = CalendarEmployeeAssignment.objects.create(
                calendar=target,
                employee=source_assignment.employee,
                operational_category=source_assignment.operational_category,
                work_pattern=source_assignment.work_pattern,
                rotation_group=source_assignment.rotation_group,
                shift_type=source_assignment.shift_type,
                five_two_off_days=source_assignment.five_two_off_days,
                default_position=source_assignment.default_position,
                start_date=source_assignment.start_date.replace(year=target.year, month=target.month, day=1),
                end_date=None,
                display_order=source_assignment.display_order,
                notes=source_assignment.notes or "",
                created_by=user,
                updated_by=user,
            )
            assignment_map[source_assignment.id] = cloned
        return assignment_map

    def _clone_cells_conservative(self, source, target, assignment_map, copy_base_cells):
        if not copy_base_cells or not assignment_map:
            return 0

        user = self._actor()
        created_count = 0
        source_cells = source.day_cells.select_related("attendance_status").order_by("assignment_id", "date")
        for source_cell in source_cells:
            target_assignment = assignment_map.get(source_cell.assignment_id)
            if not target_assignment:
                continue

            # Conservative copy: skip absences and manually overridden / exception-like entries.
            if source_cell.attendance_status and source_cell.attendance_status.is_absence:
                continue
            if source_cell.memo or source_cell.time_note or source_cell.manual_time_override:
                continue

            if source_cell.date.day > 28:
                try:
                    target_date = source_cell.date.replace(year=target.year, month=target.month)
                except ValueError:
                    continue
            else:
                target_date = source_cell.date.replace(year=target.year, month=target.month)

            CalendarDayCell.objects.update_or_create(
                assignment=target_assignment,
                date=target_date,
                defaults={
                    "calendar": target,
                    "position": source_cell.position,
                    "attendance_status": source_cell.attendance_status,
                    "work_time_code": source_cell.work_time_code,
                    "operational_code": source_cell.operational_code,
                    "scheduled_regular_minutes": source_cell.scheduled_regular_minutes,
                    "scheduled_overtime_minutes": source_cell.scheduled_overtime_minutes,
                    "actual_work_minutes": 0,
                    "actual_overtime_minutes": 0,
                    "start_time": source_cell.start_time,
                    "end_time": source_cell.end_time,
                    "break_minutes": source_cell.break_minutes,
                    "crosses_midnight": source_cell.crosses_midnight,
                    "manual_time_override": False,
                    "leave_time": None,
                    "time_note": "",
                    "overtime_minutes": 0,
                    "memo": "",
                    "raw_value": "",
                    "created_by": user,
                    "updated_by": user,
                },
            )
            created_count += 1
        return created_count

    def _clone_assignments_to_template(self, calendar, template):
        user = self._actor()
        assignment_map = {}
        source_assignments = calendar.assignments.order_by("display_order", "employee_id", "id")
        for source_assignment in source_assignments:
            cloned = OperationCalendarTemplateAssignment.objects.create(
                template=template,
                employee=source_assignment.employee,
                operational_category=source_assignment.operational_category,
                work_pattern=source_assignment.work_pattern,
                rotation_group=source_assignment.rotation_group,
                shift_type=source_assignment.shift_type,
                five_two_off_days=source_assignment.five_two_off_days,
                default_position=source_assignment.default_position,
                display_order=source_assignment.display_order,
                created_by=user,
                updated_by=user,
            )
            assignment_map[source_assignment.id] = cloned
        return assignment_map

    def _clone_cells_to_template_conservative(self, calendar, template, assignment_map, include_base_cells):
        if not include_base_cells or not assignment_map:
            return 0
        user = self._actor()
        created_count = 0
        source_cells = calendar.day_cells.select_related("attendance_status").order_by("assignment_id", "date")
        for source_cell in source_cells:
            target_assignment = assignment_map.get(source_cell.assignment_id)
            if not target_assignment:
                continue
            if source_cell.attendance_status and source_cell.attendance_status.is_absence:
                continue
            if source_cell.memo or source_cell.time_note or source_cell.manual_time_override:
                continue
            OperationCalendarTemplateCell.objects.update_or_create(
                template=template,
                template_assignment=target_assignment,
                day=source_cell.date.day,
                defaults={
                    "position": source_cell.position,
                    "attendance_status": source_cell.attendance_status,
                    "work_time_code": source_cell.work_time_code,
                    "operational_code": source_cell.operational_code,
                    "raw_value": source_cell.raw_value or "",
                    "created_by": user,
                    "updated_by": user,
                },
            )
            created_count += 1
        return created_count

    def _clone_template_assignments_to_calendar(self, template, calendar):
        user = self._actor()
        assignment_map = {}
        source_assignments = template.assignments.order_by("display_order", "employee_id", "id")
        for source_assignment in source_assignments:
            cloned = CalendarEmployeeAssignment.objects.create(
                calendar=calendar,
                employee=source_assignment.employee,
                operational_category=source_assignment.operational_category,
                work_pattern=source_assignment.work_pattern,
                rotation_group=source_assignment.rotation_group,
                shift_type=source_assignment.shift_type,
                five_two_off_days=source_assignment.five_two_off_days,
                default_position=source_assignment.default_position,
                start_date=datetime(calendar.year, calendar.month, 1).date(),
                end_date=None,
                display_order=source_assignment.display_order,
                notes="",
                created_by=user,
                updated_by=user,
            )
            assignment_map[source_assignment.id] = cloned
        return assignment_map

    def _clone_template_cells_to_calendar(self, template, calendar, assignment_map):
        if not assignment_map:
            return 0
        user = self._actor()
        created_count = 0
        template_cells = template.cells.select_related("template_assignment").order_by("template_assignment_id", "day")
        for template_cell in template_cells:
            target_assignment = assignment_map.get(template_cell.template_assignment_id)
            if not target_assignment:
                continue
            try:
                target_date = datetime(calendar.year, calendar.month, template_cell.day).date()
            except ValueError:
                continue
            CalendarDayCell.objects.update_or_create(
                assignment=target_assignment,
                date=target_date,
                defaults={
                    "calendar": calendar,
                    "position": template_cell.position,
                    "attendance_status": template_cell.attendance_status,
                    "work_time_code": template_cell.work_time_code,
                    "operational_code": template_cell.operational_code,
                    "raw_value": template_cell.raw_value or "",
                    "created_by": user,
                    "updated_by": user,
                },
            )
            created_count += 1
        return created_count

    @action(detail=True, methods=["get", "post"])
    def requirements(self, request, pk=None):
        calendar = self.get_object()

        if request.method == "GET":
            queryset = calendar.position_requirements.select_related("position").order_by("date", "position__code")
            serializer = PositionDailyRequirementSerializer(queryset, many=True)
            return Response(serializer.data)

        serializer = PositionDailyRequirementSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        user = self._actor()
        requirement = serializer.save(calendar=calendar, created_by=user, updated_by=user)
        return Response(
            PositionDailyRequirementSerializer(requirement).data,
            status=status.HTTP_201_CREATED,
        )

    @action(detail=True, methods=["patch"], url_path=r"requirements/(?P<requirement_id>\d+)")
    def requirement_detail(self, request, pk=None, requirement_id=None):
        calendar = self.get_object()
        requirement = get_object_or_404(PositionDailyRequirement, pk=requirement_id, calendar=calendar)
        serializer = PositionDailyRequirementSerializer(requirement, data=request.data, partial=True)
        serializer.is_valid(raise_exception=True)
        serializer.save(updated_by=self._actor())
        return Response(serializer.data)

    @action(detail=True, methods=["post"], url_path="requirements/replicate")
    def replicate_requirements(self, request, pk=None):
        calendar = self.get_object()
        try:
            position_id = int(request.data.get("position"))
            required_headcount = int(request.data.get("required_headcount", 0))
        except (TypeError, ValueError):
            return Response({"detail": "position e required_headcount são obrigatórios."}, status=status.HTTP_400_BAD_REQUEST)

        mode = str(request.data.get("mode") or "remaining").strip().lower()
        base_date_raw = request.data.get("date")
        notes = str(request.data.get("notes") or "")
        weekdays_only = bool(request.data.get("weekdays_only", False))
        if not base_date_raw:
            return Response({"detail": "date é obrigatório."}, status=status.HTTP_400_BAD_REQUEST)

        try:
            base_date = datetime.fromisoformat(str(base_date_raw)).date()
        except ValueError:
            return Response({"detail": "date inválida."}, status=status.HTTP_400_BAD_REQUEST)

        if base_date.year != calendar.year or base_date.month != calendar.month:
            return Response({"detail": "A data base deve pertencer ao mês do calendário."}, status=status.HTTP_400_BAD_REQUEST)

        _, total_days = monthrange(calendar.year, calendar.month)
        month_start = datetime(calendar.year, calendar.month, 1).date()
        month_end = datetime(calendar.year, calendar.month, total_days).date()
        if mode not in {"remaining", "all"}:
            return Response({"detail": "mode inválido. Use remaining ou all."}, status=status.HTTP_400_BAD_REQUEST)

        start_date = base_date if mode == "remaining" else month_start
        target_dates = []
        current = start_date
        while current <= month_end:
            if not weekdays_only or current.weekday() < 5:
                target_dates.append(current)
            current += timedelta(days=1)

        position = get_object_or_404(OperationalPosition, pk=position_id)
        user = self._actor()
        created = 0
        updated = 0
        with transaction.atomic():
            for current_date in target_dates:
                obj, was_created = PositionDailyRequirement.objects.update_or_create(
                    calendar=calendar,
                    position=position,
                    date=current_date,
                    defaults={
                        "required_headcount": required_headcount,
                        "notes": notes,
                        "updated_by": user,
                        "created_by": user,
                    },
                )
                if was_created:
                    created += 1
                else:
                    updated += 1

        return Response(
            {
                "created": created,
                "updated": updated,
                "affected_days": len(target_dates),
                "mode": mode,
                "weekdays_only": weekdays_only,
            }
        )

    @action(detail=True, methods=["get"])
    def summary(self, request, pk=None):
        calendar = self.get_object()
        requirements = calendar.position_requirements.select_related("position").order_by("date", "position__code")

        requirement_map = {
            (requirement.date, requirement.position_id): {
                "date": requirement.date,
                "position": requirement.position,
                "requirement_id": requirement.id,
                "required_headcount": requirement.required_headcount,
            }
            for requirement in requirements
        }

        assigned_counts = (
            CalendarDayCell.objects.filter(
                calendar=calendar,
                position__isnull=False,
                attendance_status__is_working_day=True,
            )
            .values("date", "position_id")
            .annotate(assigned_headcount=Count("id"))
        )

        for row in assigned_counts:
            key = (row["date"], row["position_id"])
            if key not in requirement_map:
                position = OperationalPosition.objects.get(pk=row["position_id"])
                requirement_map[key] = {
                    "date": row["date"],
                    "position": position,
                    "required_headcount": 0,
                }
            requirement_map[key]["assigned_headcount"] = row["assigned_headcount"]

        summary = []
        for item in requirement_map.values():
            assigned_headcount = item.get("assigned_headcount", 0)
            required_headcount = item["required_headcount"]
            position = item["position"]
            summary.append(
                {
                    "date": item["date"],
                    "position": position.id,
                    "requirement_id": item.get("requirement_id"),
                    "position_code": position.code,
                    "position_name_pt": position.name_pt,
                    "position_name_jp": position.name_jp,
                    "required_headcount": required_headcount,
                    "assigned_headcount": assigned_headcount,
                    "difference": assigned_headcount - required_headcount,
                }
            )

        summary.sort(key=lambda item: (item["date"], item["position_code"]))
        return Response(summary)

    def _calendar_cells(self, calendar):
        queryset = calendar.day_cells.select_related(
            "assignment",
            "assignment__employee",
            "position",
            "attendance_status",
            "work_time_code",
            "operational_code",
        )
        date_from = self.request.query_params.get("date_from")
        date_to = self.request.query_params.get("date_to")
        if date_from:
            queryset = queryset.filter(date__gte=date_from)
        if date_to:
            queryset = queryset.filter(date__lte=date_to)
        return queryset.order_by("assignment__display_order", "date")

    def _snapshot_cell(self, cell):
        if not cell:
            return None
        return {
            "position": cell.position_id,
            "attendance_status": cell.attendance_status_id,
            "work_time_code": cell.work_time_code_id,
            "operational_code": cell.operational_code_id,
            "raw_value": cell.raw_value or "",
            "overtime_minutes": cell.overtime_minutes or 0,
            "memo": cell.memo or "",
        }

    def _normalize_history_source(self, raw):
        normalized = str(raw or "").strip().lower().replace("-", "_")
        aliases = {
            "inline_edit": OperationCalendarHistory.Source.INLINE_EDIT,
            "batch_apply": OperationCalendarHistory.Source.QUICK_APPLY,
            "batch_status": OperationCalendarHistory.Source.QUICK_APPLY,
            "batch_op_code": OperationCalendarHistory.Source.QUICK_APPLY,
            "batch_clear": OperationCalendarHistory.Source.QUICK_APPLY,
            "paste": OperationCalendarHistory.Source.PASTE,
            "fill_handle": OperationCalendarHistory.Source.FILL_HANDLE,
            "pattern_4x2": OperationCalendarHistory.Source.PATTERN_4X2,
            "template": OperationCalendarHistory.Source.TEMPLATE,
            "month_duplication": OperationCalendarHistory.Source.MONTH_DUPLICATION,
            "next_month_generation": OperationCalendarHistory.Source.NEXT_MONTH_GENERATION,
        }
        return aliases.get(normalized, OperationCalendarHistory.Source.SYSTEM)

    def _log_cell_history(self, calendar, assignment, cell_date, source, old_value, new_value, metadata=None):
        try:
            OperationCalendarHistory.objects.create(
                calendar=calendar,
                assignment=assignment,
                cell_date=cell_date,
                source=source or OperationCalendarHistory.Source.SYSTEM,
                old_value=old_value,
                new_value=new_value,
                metadata=metadata or {},
                created_by=self._actor(),
                updated_by=self._actor(),
            )
        except Exception:
            return

    def _log_operation_history(self, calendar, source, metadata=None):
        try:
            OperationCalendarHistory.objects.create(
                calendar=calendar,
                source=source,
                metadata=metadata or {},
                created_by=self._actor(),
                updated_by=self._actor(),
            )
        except Exception:
            return

    def _paste_tsv(self, calendar, assignments, start_date, tsv, source=OperationCalendarHistory.Source.PASTE):
        parser_context = build_calendar_cell_parser_context(calendar)

        rows = list(csv.reader(StringIO(tsv), delimiter="\t"))
        if rows and rows[-1] == [""]:
            rows.pop()

        created = 0
        updated = 0
        affected_cells = []
        unrecognized_values = []
        actor = self._actor()

        with transaction.atomic():
            for row_index, row_values in enumerate(rows):
                if row_index >= len(assignments):
                    break

                assignment = assignments[row_index]
                for column_index, raw_value in enumerate(row_values):
                    cell_date = start_date + timedelta(days=column_index)
                    if cell_date.year != calendar.year or cell_date.month != calendar.month:
                        break

                    value = raw_value.strip()
                    parsed = parse_calendar_cell_value(value, parser_context)
                    cell, was_created = CalendarDayCell.objects.get_or_create(
                        assignment=assignment,
                        date=cell_date,
                        defaults={
                            "calendar": calendar,
                            "created_by": actor,
                        },
                    )
                    old_snapshot = None if was_created else self._snapshot_cell(cell)

                    cell.calendar = calendar
                    cell.raw_value = value
                    cell.position = parsed.position
                    cell.attendance_status = parsed.attendance_status
                    cell.work_time_code = parsed.work_time_code
                    cell.operational_code = parsed.operational_code
                    cell.memo = parsed.memo
                    cell.updated_by = actor
                    calculate_cell_work_minutes(cell, persist=False)
                    cell.save()
                    self._log_cell_history(
                        calendar=calendar,
                        assignment=assignment,
                        cell_date=cell_date,
                        source=source,
                        old_value=old_snapshot,
                        new_value=self._snapshot_cell(cell),
                        metadata={"origin": "paste_endpoint"},
                    )

                    if was_created:
                        created += 1
                    else:
                        updated += 1

                    affected_cells.append(
                        {
                            "id": cell.id,
                            "assignment": assignment.id,
                            "date": cell_date,
                            "raw_value": value,
                            "created": was_created,
                        }
                    )

                    if value and not parsed.recognized:
                        unrecognized_values.append(
                            {
                                "assignment": assignment.id,
                                "date": cell_date,
                                "value": value,
                            }
                        )

        return {
            "created": created,
            "updated": updated,
            "unrecognized_values": unrecognized_values,
            "affected_cells": affected_cells,
        }
